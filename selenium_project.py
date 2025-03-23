# # 輸出中文
# from importlib import reload
# import sys
# reload(sys)
# sys.serdefaultencoding('utf-8')

#selenium
import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select

# local excel
import openpyxl as op

filepath = "School.xlsx"
try:
    workbook = op.load_workbook(filepath)
except:
    op.Workbook().save(filepath)
                       
workbook = op.load_workbook(filepath) 
sheet = workbook.worksheets[0]



chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

service = Service(executable_path=r"C:\Users\user\Desktop\chromedriver-win64\chromedriver.exe")
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.get("https://tyc1.fdhs.tyc.edu.tw/tyc1/interview_search.php") # 獲取網站
wait = WebDriverWait(driver, 3) # 設定等待時間



ExcelIndex = 1


def initial(n):
    # 搜尋大學的
    #wait = WebDriverWait(driver, 10)
    try: # 勾選要的大學
        # 防止 stale element reference: element is not attached to the page document，網頁更新兩次，而只抓到第一次的資料
        select_element = Select(wait.until(EC.visibility_of_element_located((By.XPATH, "//select[@name='s_university_value']"))))
        select_element.select_by_index(n)

    except:
        select_element = Select(wait.until(EC.visibility_of_element_located((By.XPATH, "//select[@name='s_university_value']"))))
        select_element.select_by_index(n)
    try: # 要 100 筆資料的筆資料的
        selectNumbers = Select(wait.until(EC.visibility_of_element_located((By.XPATH, "//select[@name='p_l']"))))
        selectNumbers.select_by_index(9)
    except:
        selectNumbers = Select(wait.until(EC.visibility_of_element_located((By.XPATH, "//select[@name='p_l']"))))
        selectNumbers.select_by_index(9)
    # 按查詢
    clickSearch = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@type='submit'][@value='查詢']")))
    
    clickSearch.click()

def goToQuestionBank(singleUniversityValue):
    
    #multiple_elements
    
    
    # 讓多筆資料也能 wait 等他跑完
    
    print(f"下面是該科系在此大學的序號{singleUniversityValue}") # 1 base
    #print(((int(num) - 1) % 100+1+1))
    print()
    
    # 記得要再用 ()把 By.XPATH 括號起來                                                                                                                        # n = (n-1) % 100 + 1// 後面的加一為了找位置
    wait.until(EC.visibility_of_element_located((By.XPATH, f"/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table[3]/tbody/tr[{((int(singleUniversityValue) - 1) % 100+1+1)}]/td[6]/input[18]"))).click()
    time.sleep(0.01)
    
    detail(singleUniversityValue)
    print()
    
    #leave
    back = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@type='submit'][@value='回上一頁']")))
    back.click()
   
        
def SaveToExcel(data, singleUniversityValue): # 存入 excel
    
    sheet[f'A{ExcelIndex}'] = singleUniversityValue
    for i in range(0, len(data)):
        sheet.cell(row = ExcelIndex, column=i+2).value = data[i]
    
    workbook.save(filepath)
        

# num 只是當前大學順序，不會持續增加，會有歸為 1 的一天     
def detail(singleUniversityValue): #會受網路影響，而藥用 WebDriverWait
    #wait = WebDriverWait(driver, 10)
    global ExcelIndex
    
    # 考試年度 申請學校 申請校系 面試方式 面試問題 筆試題目
    save = ['無資料','無資料','無資料','無資料','無資料', '無資料']  
    for k in range(1, 11):
        try:
            aim = wait.until(EC.visibility_of_element_located((By.XPATH, f"/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[{k}]/td[1]")))
        except:
            break
        
        if aim.text == "考試年度":
            target = wait.until(EC.visibility_of_element_located((By.XPATH, f"/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[{k}]/td[2]")))
            save[0]= target.text
        elif aim.text == "申請校系":
            # 會爆掉 中央大學 32
            target = wait.until(EC.visibility_of_element_located((By.XPATH, f"/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[{k}]/td[2]")))
            if (len(target.text.split()) < 2):
                ExcelIndex += 1
                return
            save[1], save[2] = target.text.split()
            #save.append(target.text)
        elif aim.text == "面試方式":
            target = wait.until(EC.visibility_of_element_located((By.XPATH, f"/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[{k}]/td[2]")))
            save[3] = target.text
        elif aim.text == "面試問題":
            target = wait.until(EC.visibility_of_element_located((By.XPATH, f"/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[{k}]/td[2]")))
            save[4] = target.text
        elif aim.text == "筆試題目":
            target = wait.until(EC.visibility_of_element_located((By.XPATH, f"/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr[{k}]/td[2]")))
            save[5] = target.text
        
            break
    for _ in range(len(save)):
        print(save[_])
    print("\n資料完整獲取!")      
    SaveToExcel(save, singleUniversityValue)
    ExcelIndex += 1
    return

def main():
    global ExcelIndex
    print("Loading...")
    selectUniversity = WebDriverWait(driver, 3).until( # 申請大學清單旁邊的按鈕
        EC.presence_of_element_located((By.NAME, 's_university'))
    )
    #selectUniversity = driver.find_element(By.NAME,'s_university')
    selectUniversity.click()
    select = Select(driver.find_element(By.NAME,'s_university_value')) # 選大學表單清單
    print("Successful!")

 
    '''
    維修方式
    1 是中山大學，其他則照表單順序
    記得初始化 ExcelIndex 不然會跑錯
    '''
    print("啟動 BUG 維修模式嗎?")
    ask = input("(y/n)\n")
    if(ask == 'y'):
        #ExcelIndex
        #initial(n)
        #page
        #singleUniversityValue
        ExcelIndex = int(input("輸入當前要儲存 Excel 序號\n"))
        initialNumber = input("輸入大學的編碼\n")
        pageNumber = int(input("輸入要輸入的資料在第幾頁\n")) - 1
        selectUniversityValue = int(input("輸入要輸入的的該大學科系的序號\n"))
          
        initial(initialNumber) # 大學選擇清單序號序號
        totalData = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/center/font/b"))
        ).text

        for i in range(pageNumber): # click 換下一頁 
            WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='submit'][@value='下一頁']"))
            ).click()
            time.sleep(0.1)
            
        for singleUniversityValue in range(selectUniversityValue, int(totalData) + 1):
            print(f'目前是第 {ExcelIndex} 筆資料')     
            goToQuestionBank(singleUniversityValue)
            if singleUniversityValue % 100 == 0:
                nextPage = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='submit'][@value='下一頁']"))
                )
                nextPage.click()
        print("Bug 維修完畢，已確實填入該大學所有校系\n")
        driver.quit()
    

    universityIndex = int(input("請輸入大學的排序位置\n"))
    
    for i in range(universityIndex, len(select.options)):
        
        initial(i) # 共有 i 筆大學

        totalData = WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/center/font/b"))
        ).text
        #totalData = driver.find_element(By.XPATH, "//center/font").text
        print(f'該大學總共有{totalData}資料')
        #(n-1) % 100 + 1
        
        # 第 num 筆資料
        #singleUniversityValue
        print(ExcelIndex)
        for singleUniversityValue in range(1, int(totalData) + 1):
            print(f'目前是 Excel 表格的第 {ExcelIndex} 筆資料')     
            goToQuestionBank(singleUniversityValue)
            if singleUniversityValue % 100 == 0:
                nextPage = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='submit'][@value='下一頁']"))
                )
                nextPage.click()
        print("\nfinish!\n")
        time.sleep(10)

if __name__ == '__main__':
    main()